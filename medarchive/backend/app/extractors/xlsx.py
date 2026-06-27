"""XLSX/XLS extractor (brief §7). Iterates ALL sheets, detects the header row
(which may not be row 1), maps service + one/two price columns, skips empties."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.extractors.base import (
    ExtractedRow,
    ExtractionResult,
    classify_columns,
    looks_like_header,
    register,
)
from app.models.enums import FileFormat


@register(FileFormat.xlsx)
class XlsxExtractor:
    def extract(self, path: Path) -> ExtractionResult:
        wb = load_workbook(path, read_only=True, data_only=True)
        rows: list[ExtractedRow] = []
        text_lines: list[str] = []
        notes: list[str] = []

        for ws in wb.worksheets:
            sheet_rows = list(ws.iter_rows(values_only=True))
            if not sheet_rows:
                continue
            text_lines.append(f"# sheet: {ws.title}")

            header_idx = self._find_header(sheet_rows)
            if header_idx is None:
                notes.append(f"sheet '{ws.title}': no header detected, skipped")
                continue
            cols = classify_columns(list(sheet_rows[header_idx]))
            if cols["name"] is None:
                cols["name"] = 0  # fall back to first column
            notes.append(
                f"sheet '{ws.title}': header@row{header_idx + 1} cols={cols}"
            )

            for raw in sheet_rows[header_idx + 1:]:
                if raw is None or all(c is None for c in raw):
                    continue
                line = " | ".join("" if c is None else str(c) for c in raw)
                text_lines.append(line)
                name = self._cell(raw, cols["name"])
                if not name or self._is_total(name):
                    continue
                rows.append(
                    ExtractedRow(
                        service_name_raw=name,
                        price_resident=self._cell(raw, cols["resident"]),
                        price_nonresident=self._cell(raw, cols["nonresident"]),
                        specialty_hint=ws.title if ws.title.lower() not in ("прайс", "sheet1") else None,
                        source_fragment=line[:200],
                    )
                )
        wb.close()
        return ExtractionResult(rows=rows, raw_text="\n".join(text_lines), log="; ".join(notes))

    @staticmethod
    def _find_header(sheet_rows: list, scan: int = 15) -> int | None:
        for i, row in enumerate(sheet_rows[:scan]):
            if row and looks_like_header(list(row)):
                return i
        return None

    @staticmethod
    def _cell(row: tuple, idx: int | None):
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return str(v).strip() if isinstance(v, str) else v

    @staticmethod
    def _is_total(name: str) -> bool:
        low = str(name).strip().lower()
        return low.startswith(("итого", "всего", "total", "сумма"))
