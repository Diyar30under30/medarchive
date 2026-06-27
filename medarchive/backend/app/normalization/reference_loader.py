"""Load the canonical directory into the `services` table (ТЗ §2.2, §3.4).

Schema-flexible: handles BOTH the organizer-spec directory
(`service_id, service_name, synonyms, category, icd_code`) AND the real
`Справочник услуг.xlsx` layout (`ID, Специальность, Code, Name_ru,
TarificatrCode`), in XLSX or JSON. Columns are detected by name (case-
insensitive), a provided `service_id` is used when present (else a deterministic
UUID5 is generated), and a provided `synonyms` array / `icd_code` are loaded.
Idempotent: re-running upserts and MERGES provided synonyms with operator-learned
ones (never loses the learning loop).
"""
from __future__ import annotations

import json
import logging
import re
import uuid
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import settings
from app.models.service import Service
from app.normalization.embeddings import get_embedding_provider

log = logging.getLogger("medarchive.reference")

# Stable namespace so generated UUID5s are reproducible across runs/machines.
_NS = uuid.UUID("d3f1c0de-0000-4000-8000-000000000001")

# field → candidate column names (lowercased, exact match after strip).
# Order matters only for documentation; matching is by membership.
_FIELD_ALIASES: dict[str, list[str]] = {
    "service_id": ["service_id", "serviceid", "service id", "id_услуги", "uuid"],
    "service_name": ["service_name", "name_ru", "name", "наименование", "услуга", "service", "название"],
    "category": ["category", "специальность", "категория", "раздел"],
    "synonyms": ["synonyms", "синонимы", "synonym", "альтернативные названия"],
    "icd_code": ["icd_code", "icd", "icd10", "мкб", "код мкб"],
    "source_code": ["code", "source_code", "код"],
    "tariff_code": ["tarificatrcode", "tariff_code", "тариф", "тарификатор"],
}


def service_uuid(category: str, source_code: str, name: str) -> str:
    return str(uuid.uuid5(_NS, f"{category}|{source_code}|{name}".strip()))


def _detect_columns(headers: list[str]) -> dict[str, str | None]:
    norm = {h: str(h).strip().lower() for h in headers if h is not None}
    cols: dict[str, str | None] = {}
    for field, aliases in _FIELD_ALIASES.items():
        match = next((h for h, low in norm.items() if low in aliases), None)
        cols[field] = match
    return cols


def _parse_synonyms(value) -> list[str]:
    if value is None or value == "":
        return []
    if isinstance(value, (list, tuple)):
        return [str(x).strip() for x in value if str(x).strip()]
    s = str(value).strip()
    if not s:
        return []
    try:  # JSON array stored as a string
        j = json.loads(s)
        if isinstance(j, list):
            return [str(x).strip() for x in j if str(x).strip()]
    except (json.JSONDecodeError, ValueError):
        pass
    return [p.strip() for p in re.split(r"[;,|/]", s) if p.strip()]


def _clean(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _normalize_row(raw: dict, cols: dict[str, str | None]) -> dict | None:
    def get(field: str):
        col = cols.get(field)
        return raw.get(col) if col else None

    name = _clean(get("service_name"))
    if not name:
        return None
    category = _clean(get("category")) or "Без категории"
    source_code = _clean(get("source_code")) or ""
    provided_id = _clean(get("service_id"))
    return {
        "service_id": provided_id or service_uuid(category, source_code, name),
        "service_name": name,
        "category": category,
        "source_code": source_code or None,
        "tariff_code": _clean(get("tariff_code")),
        "icd_code": _clean(get("icd_code")),
        "synonyms": _parse_synonyms(get("synonyms")),
    }


def _read_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    # Header may not be row 1 — find the first row that looks like field headers.
    all_aliases = {a for al in _FIELD_ALIASES.values() for a in al}
    header_idx = 0
    for i, r in enumerate(rows[:10]):
        cells = [str(c).strip().lower() for c in r if c is not None]
        if sum(c in all_aliases for c in cells) >= 2:
            header_idx = i
            break
    headers = [str(h).strip() if h is not None else "" for h in rows[header_idx]]
    out: list[dict] = []
    for raw in rows[header_idx + 1:]:
        if raw is None or all(c is None for c in raw):
            continue
        out.append({headers[i]: raw[i] for i in range(min(len(headers), len(raw)))})
    return out


def _read_json(path: Path) -> list[dict]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        data = data.get("services") or data.get("rows") or data.get("items") or []
    return list(data)


def find_reference_file() -> Path | None:
    ref_dir = settings.reference_path
    if not ref_dir.exists():
        return None
    for pattern in ("*.json", "Справочник услуг.xlsx", "*.xlsx", "*.xls"):
        hits = sorted(ref_dir.glob(pattern))
        if hits:
            return hits[0]
    return None


def load_reference(db: Session, path: Path | None = None) -> dict:
    path = path or find_reference_file()
    if path is None or not path.exists():
        log.warning("No reference file found in %s", settings.reference_path)
        return {"loaded": 0, "updated": 0, "skipped": 0, "path": None}

    rows = _read_json(path) if path.suffix.lower() == ".json" else _read_xlsx(path)
    cols = _detect_columns(list(rows[0].keys())) if rows else {}
    log.info("Reference columns detected: %s", {k: v for k, v in cols.items() if v})

    provider = get_embedding_provider()
    normalized = [n for n in (_normalize_row(r, cols) for r in rows) if n]
    embeddings: list[list[float]] | None = None
    if provider.available and normalized:
        embeddings = provider.encode_batch([n["service_name"] for n in normalized])

    created = updated = 0
    with_synonyms = 0
    existing = {s.service_id: s for s in db.scalars(select(Service)).all()}
    for i, n in enumerate(normalized):
        emb = embeddings[i] if embeddings else None
        if n["synonyms"]:
            with_synonyms += 1
        svc = existing.get(n["service_id"])
        if svc is None:
            svc = Service(
                service_id=n["service_id"],
                service_name=n["service_name"],
                category=n["category"],
                source_code=n["source_code"],
                tariff_code=n["tariff_code"],
                icd_code=n["icd_code"],
                synonyms=sorted(set(n["synonyms"])),
                embedding=emb,
                is_active=True,
            )
            db.add(svc)
            created += 1
        else:
            svc.service_name = n["service_name"]
            svc.category = n["category"]
            svc.source_code = n["source_code"]
            svc.tariff_code = n["tariff_code"]
            if n["icd_code"]:
                svc.icd_code = n["icd_code"]
            # Merge provided synonyms with operator-learned ones (keep the loop).
            if n["synonyms"]:
                svc.synonyms = sorted(set(svc.synonyms or []) | set(n["synonyms"]))
            if emb is not None:
                svc.embedding = emb
            updated += 1
    db.commit()

    result = {
        "loaded": created,
        "updated": updated,
        "skipped": len(rows) - len(normalized),
        "total": len(normalized),
        "with_provided_synonyms": with_synonyms,
        "embeddings": bool(embeddings),
        "columns": {k: v for k, v in cols.items() if v},
        "path": str(path),
    }
    log.info("Reference load: %s", result)
    return result
